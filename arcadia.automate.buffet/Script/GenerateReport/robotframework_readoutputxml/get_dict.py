from new_ReaderRobotFramework import ReaderRobotFramework
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from styleframe import StyleFrame, Styler, utils
from datetime import datetime
import os   
import re


# FEATURE = ['ClientCredential', 'LDAPallOU', 'ChangePassword', 'EmailOTP', 'FBBOTP', 'ValidateToken']
FEATURE = ['LDAPallOU']
PATH_OUTPUT_XML_FILE: str = r'../../../../admd.automate.robot.testing/TestSuite/Result/output.xml'
# PATH_OUTPUT_IMAGE: str = r'C:\Users\mc\Arcadia\Andromeda\admd.automate.robot.testing\Result'
PATH_OUTPUT_REPORT: str = r'report_excel.xlsx'
REPORT_HEADER: dict = {
    'Test Case ID': 25.0,               # testcase_id
    'Test Case Description': 50.0,      # testcase_only_name
    'Pre-Requisite': 50.0,              # documentation
    'Test Step Description': 50.0,      # step_keyword
    'Expected Result': 25.0,            #
    'Actual Result': 50.0,              #
    'Screenshot': 70.0,                 #
    'Date': 25.0,                       # date_time
    'Status': 10.0,                     # test_result
    'Remark': 50.0                      # keyword_fail, msg_error
}
SET_HEADER_ROW_HEIGHT = 30
SET_ROW_HEIGHT = 200
SIZE_IMAGE_WIDTH = 480
SIZE_IMAGE_HEIGHT = 245


def read_output_xml(feature_name: str):
    keyword_regex_ignore = ['Run Keyword And Ignore Error']
    reader = ReaderRobotFramework(PATH_OUTPUT_XML_FILE, feature_name, keyword_regex_ignore)
    robot_result: dict = reader.get_robot_output_result()

    feature_data: dict = robot_result[feature_name]
    testcase: list = feature_data['TestcaseDetail']
    dataframe: list = []

    for testcase_detail in testcase:
        testcase_name: str = testcase_detail['testcase_name']
        testcase_name_split = testcase_name.split(' ', 1)
        testcase_id: str = testcase_name_split[0]
        testcase_only_name: str = testcase_name_split[1]
        documentation: str = testcase_detail['documentation']
        doc_line = documentation.split('\n')
        expected_result: str = ''
        set_expected_result: bool = False
        actual_result: str = ''
        set_actual_result: bool = False
        documentation_format: str = ''
        for line in doc_line:
            # check text is exist
            if line:
                if line.startswith('Owner') or line.startswith('owner'):
                    continue
                elif line.startswith('***Expected Result***') or set_expected_result:
                    if line.startswith('***Provisioning data***'):
                        set_expected_result = False
                        documentation_format += f'{line}\r\n'
                        continue
                    if not set_expected_result:
                        set_expected_result = True
                        continue
                    expected_result += f'{line}\r\n'
                elif line.startswith('***Actual Result***') or set_actual_result:
                    if not set_actual_result:
                        set_actual_result = True
                        continue
                    actual_result += f'{line}\r\n'
                else:
                    documentation_format += f'{line}\r\n'

        tags: str = testcase_detail['tags']
        step_keyword: str = testcase_detail['step_keyword']
        test_result: str = testcase_detail['test_result']
        keyword_fail: str = testcase_detail['keyword_fail']
        date_time: str = testcase_detail['date_time']
        date_time_format = datetime.strptime(date_time, '%Y%m%d  %H:%M:%S.%f')
        msg_error: str = testcase_detail['msg_error']
        detail: list = [
                    testcase_id,
                    testcase_only_name,
                    documentation_format,
                    step_keyword,
                    expected_result,
                    actual_result,
                    '',
                    f'{date_time_format:%d %b %Y %H:%M:%S}',
                    test_result,
                    f'{keyword_fail}\n{msg_error}'
                    ]
        dataframe.append(detail)

    return dataframe


def test_excel():
    excel_writer = StyleFrame.ExcelWriter(PATH_OUTPUT_REPORT)

    for feature_name in FEATURE:
        dataframe = read_output_xml(feature_name)
        headers: list = list(REPORT_HEADER.keys())
        df = pd.DataFrame(dataframe, columns=headers)

        # Style
        default_style = Styler(
                            horizontal_alignment='left',
                            vertical_alignment='top',
                            font=utils.fonts.calibri,   # Add font format
                            font_size=13
                            )
        sf = StyleFrame(df, styler_obj=default_style)

        # Header style
        header_style = Styler(
                            horizontal_alignment='center',
                            vertical_alignment='center',
                            bg_color=utils.colors.dark_green,
                            bold=True,
                            font=utils.fonts.calibri,   # Add font format
                            font_size=15
                            )
        sf.apply_headers_style(styler_obj=header_style)

        # Header style - Date, Status
        sf.apply_column_style(cols_to_style=['Date', 'Status'],
                            styler_obj=Styler(
                                            horizontal_alignment='center',
                                            vertical_alignment='top'
                                            )
                            )

        # Set Width
        width_set: dict = {}
        list_header_set: list = []
        index_width = 0
        for index_width, width in enumerate(REPORT_HEADER):
            if width in width_set:
                list_header_set = width_set[width]
            else:
                list_header_set: list = []
            list_header_set.append(headers[index_width])
            width_set[width] = list_header_set

        width_keys: list = list(width_set.keys())
        for index_key_width, column_header_set in enumerate(width_set):
            sf = sf.set_column_width(columns=[column_header_set], width=REPORT_HEADER[width_keys[index_key_width]])

        # Set Height
        sf.set_row_height(rows=sf.row_indexes, height=SET_ROW_HEIGHT)
        sf.set_row_height(rows=1, height=SET_HEADER_ROW_HEIGHT)
        
        sf.to_excel(excel_writer, sheet_name=feature_name)

    excel_writer.save()


def insert_image():
    workbook = load_workbook(PATH_OUTPUT_REPORT)

    for feature_name in FEATURE:
        worksheet = workbook[feature_name]

        image_insert: dict = {}
        for file in os.listdir(PATH_OUTPUT_IMAGE):
            if file.endswith(".png"):
                match = re.search('\w+_\w\d+_\d_\d_\d{3}', file)
                if match:
                    # check image capture screen latest
                    if match.group() in image_insert:
                        file_name = os.path.splitext(image_insert[match.group()])[0]
                        new_file_name = os.path.splitext(file)[0]
                        if int(file_name[-1]) > int(new_file_name[-1]):
                            file = file_name
                    image_insert[match.group()] = file

        # Row of data (testcase detail)
        for row in range(2, len(worksheet['A'])+1):
            testcase_id:str = worksheet.cell(row=row, column=1).value
            if testcase_id in image_insert:

                img = Image(f'{PATH_OUTPUT_IMAGE}\{image_insert[testcase_id]}')
                img.width = SIZE_IMAGE_WIDTH
                img.height = SIZE_IMAGE_HEIGHT
                img.anchor = f'G{row}'
                worksheet.add_image(img)

    workbook.save(PATH_OUTPUT_REPORT)


test_excel()
# insert_image()
